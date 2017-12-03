import openpyxl as px
import pandas as pd
from pandas import ExcelWriter
from pandasql import sqldf

url = "DD.xlsx"
xl = pd.read_excel(url, "Sheet1", 0)
pysqldf = lambda q: sqldf(q, globals())
# Fill in Data .
wb = px.load_workbook(url)
ws = wb.get_sheet_by_name("Sheet1")
sheet = wb.active

def replace_module():

    def ID():



        nRow = 1
        nlRow = ws.max_row + 1
        nlColumn = ws.max_column



        while True:
            IDinput = input("Input ID of Student to be replaced (or 0 to exit): ")
            while not IDinput.isdigit():
                print('Invalid Input')
                IDinput = input("Input ID of Student to be replaced: ")
            if IDinput.isdigit():
                for row in range(nRow, nlRow):
                    id = ws['A' + str(row)].value
                    status = ws['L' + str(row)].value
                    if str(IDinput) == id and status=='A':

                        FName = ws['B' + str(row)].value
                        LName = ws['C' + str(row)].value
                        gender = (ws['E' + str(row)].value)
                        talent = ws['M' + str(row)].value
                        instrument = ws['N' + str(row)].value
                        dorm = ws['O' + str(row)].value
                        band = ws['P' + str(row)].value
                        print(id, FName+' '+LName,gender, talent,instrument,dorm,band,status )

                        confirmInput = input('\n'.join(['Confirm: ', '1: Yes', '2: No', '']))
                        while confirmInput not in ['1', '2']:
                            print('Invalid Input')
                            confirmInput = input('\n'.join(['Confirm: ', '1: Yes', '2: No', '']))
                        if confirmInput == '1':

                            searchid = input('\n'.join(['Insert Waiting list Student: ', '1: By ID', '2: View List', '3: Cancel','']))
                            while searchid not in ['1', '2','3']:
                                print('Invalid Input')
                                searchid = input('\n'.join(['Insert Waiting list Student: ', '1: By ID', '2: View List', '3: Cancel','']))
                            if searchid == '1':

                                ID2input = input("Enter the numeric ID of the waiting list student (or 0 to exit): ")
                                while not ID2input.isdigit():
                                    print('Invalid Input')
                                    ID2input = input("Enter the numeric ID of the waiting list student (or 0 to exit): ")
                                if ID2input!='0':
                                    for crow in range(nRow, nlRow):
                                        id2 = ws['A' + str(crow)].value
                                        status2 = ws['L' + str(crow)].value
                                        if str(ID2input) == id2 and status2 == 'W':

                                            FName2 = ws['B' + str(crow)].value
                                            LName2 = ws['C' + str(crow)].value
                                            gender2 = (ws['E' + str(crow)].value)
                                            talent2 = ws['M' + str(crow)].value
                                            instrument2 = ws['N' + str(crow)].value
                                            Dorm2 = str(ws['O' + str(crow)].value)
                                            band2 = str(ws['P' + str(crow)].value)
                                            print("Replace Student",'\n', id, FName + ' ' + LName, gender,
                                                  talent, instrument, dorm, band, status, '\n', "with",
                                                  '\n', id2, FName2 + ' ' + LName2, gender2, talent2,
                                                  instrument2, Dorm2, band2, status2)
                                            confirm2Input = input(
                                                '\n'.join(['Confirm: ', '1: Yes', '2: No', '']))
                                            while confirm2Input not in ['1', '2']:
                                                print('Invalid Input')
                                                confirm2Input = input(
                                                    '\n'.join(['Confirm: ', '1: Yes', '2: No', '']))
                                            if confirm2Input == '1':
                                                ws['O' + str(crow)].value = dorm

                                                ws['P' + str(crow)].value = band
                                                ws['P' + str(row)].value = ''
                                                ws['O' + str(row)].value = ''
                                                ws['L' + str(crow)].value = 'A'
                                                ws['L' + str(row)].value = 'D'

                                                wb.save(url)

                                                break

                                            if confirm2Input == '2':
                                                break
                                            break


                                if ID2input == '0':
                                    break

                            if searchid == '2':
                                print(pysqldf("Select [ID Number],[First Name], [Last Name],[Gender],[Instrument],[Talent], [Status] from xl where [Status]='W' order by [ID Number]"))
                                break
                            if searchid == '3':
                                break


                            break
                        if confirmInput == '2':
                            break

                        break

            if IDinput == '0':
                break
            break

    while True:
        humanInput = input('\n'.join(
            ['Replace student by: ', '1: ID', '2: Name', '0: Go Back', '']))
        while humanInput not in ['1', '2', '0']:
            print('Invalid Input')
            humanInput = input('\n'.join(
                ['Replace student by: ', '1: ID', '2: Name', '0: Go Back', '']))
        if humanInput == '1':
            ID()
        if humanInput == '0':
            break