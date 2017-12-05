import openpyxl as px
import pandas as pd
from pandas import ExcelWriter
from pandasql import sqldf
import warnings

def First_Band_Assign():
    url = "DD.xlsx"
    xl = pd.read_excel(url, "Sheet1", 0)
    warnings.simplefilter("ignore") 
    # Fill in Data .
    wb = px.load_workbook(url)
    ws = wb.get_sheet_by_name("Sheet1")
    sheet = wb.active
    b1 = {'gender':[],'type':'1'}
    b2 = {'gender':[],'type':'1'}
    b3 = {'gender':[],'type':'1'}
    b4 = {'gender':[],'type':'1'}
    b5 = {'gender':[],'type':'2'}
    b6 = {'gender':[],'type':'2'}
    b7 = {'gender':[],'type':'2'}
    b8 = {'gender':[],'type':'2'}
    Bs1 = []
    Bs2 = []
    nRow = 1
    nlRow = ws.max_row + 1
    nlColumn = ws.max_column


    for row in range(nRow, nlRow):
        status = ws['L' + str(row)].value
        if status == 'A':
            gender = ws['E' + str(row)].value
            ins = ws['N' + str(row)].value
            talent = int(ws['M' + str(row)].value)
            if check(b1,b1['gender'],gender,ins,talent,Bs1,Bs2):
                b1[ins] = talent
                b1['gender'].append(gender)
                ws["P"+str(row)] = 1
            elif check(b2,b2['gender'],gender,ins,talent,Bs1,Bs2):
                b2[ins] = talent
                b2['gender'].append(gender)
                ws["P"+str(row)] = 2
            elif check(b3,b3['gender'],gender,ins,talent,Bs1,Bs2):
                b3[ins] = talent
                b3['gender'].append(gender)
                ws["P"+str(row)] = 3
            elif check(b4,b4['gender'],gender,ins,talent,Bs1,Bs2):
                b4[ins] = talent
                b4['gender'].append(gender)
                ws["P"+str(row)] = 4
            elif check(b5,b5['gender'],gender,ins,talent,Bs1,Bs2):
                b5[ins] = talent
                b5['gender'].append(gender)
                ws["P"+str(row)] = 5
            elif check(b6,b6['gender'],gender,ins,talent,Bs1,Bs2):
                b6[ins] = talent
                b6['gender'].append(gender)
                ws["P"+str(row)] = 6
            elif check(b7,b7['gender'],gender,ins,talent,Bs1,Bs2):
                b7[ins] = talent
                b7['gender'].append(gender)
                ws["P"+str(row)] = 7
            elif check(b8,b8['gender'],gender,ins,talent,Bs1,Bs2):
                b8[ins] = talent
                b8['gender'].append(gender)
                ws["P"+str(row)] = 8
        wb.save(url)


#band2 assign: check talent -> make choice on age -> return ID -> Assigning
def Second_Band_Assign():
    url = "DD.xlsx"
    xl = pd.read_excel(url, "Sheet1", 0)
    # Fill in Data .
    wb = px.load_workbook(url)
    ws = wb.get_sheet_by_name("Sheet1")
    sheet = wb.active

    mb1 = {'age':[],'ID':"1"}
    mb2 = {'age':[],'ID':"2"}
    mb3 = {'age':[],'ID':"3"}
    mb4 = {'age':[],'ID':"4"}

    fb1 = {'age':[],'ID':"1"}
    fb2 = {'age':[],'ID':"2"}
    fb3 = {'age':[],'ID':"3"}
    fb4 = {'age':[],'ID':"4"}
    nRow = 1
    nlRow = ws.max_row + 1
    nlColumn = ws.max_column
    for row in range(nRow, nlRow):
        status = ws['L' + str(row)].value
        if status == 'A':
            gender = ws['E' + str(row)].value
            age = int(ws['F' + str(row)].value)
            ins = ws['N' + str(row)].value
            talent = int(ws['M' + str(row)].value)
            if gender == "M":
                b_total = talentStatus(mb1,mb2,mb3,mb4,ins,talent)
                choice = ageChoice(b_total,age)
                if choice == "1":
                    ws['Q'+str(row)] = 'M1'
                    mb1['age'].append(age)
                    mb1[ins] = talent
                elif choice == "2":
                    ws['Q'+str(row)] = 'M2'
                    mb2['age'].append(age)
                    mb2[ins] = talent
                elif choice == "3":
                    ws['Q'+str(row)] = 'M3'
                    mb3['age'].append(age)
                    mb3[ins] = talent
                elif choice == "4":
                    ws['Q'+str(row)] = 'M4'
                    mb4[ins] = talent 
            elif gender == "F":
                b_total = talentStatus(fb1,fb2,fb3,fb4,ins,talent)
                choice = ageChoice(b_total,age)
                if choice == "1":
                    ws['Q'+str(row)] = 'F1'
                    fb1['age'].append(age)
                    fb1[ins] = talent
                elif choice == "2":
                    ws['Q'+str(row)] = 'F2'
                    fb2['age'].append(age)
                    fb2[ins] = talent
                elif choice == "3":
                    ws['Q'+str(row)] = 'F3'
                    fb3['age'].append(age)
                    fb3[ins] = talent
                elif choice == "4":
                    ws['Q'+str(row)] = 'F4'
                    fb4['age'].append(age)
                    fb4[ins] = talent  
        wb.save(url)
def checkBandStatus(Bs1,Bs2,band,ins,talent):
    mockBand = band
    mockBand[ins] = talent
    category = Band_Category(mockBand)
    if category == 1:
        if band not in Bs1:
            return False
        else:
            return True
    else:
        if band not in Bs2:
            return False
        else:
            return True




def Band_Category(band):
    talentList = [countTalent(band,1),countTalent(band,2),countTalent(band,3),countTalent(band,4)]
    if talentList[0] == 2 or talentList[3] == 2:
        return 1
    elif talentList[1] == 2 or talentList[2] == 2:
        return 2

def check(band,band_gender,gender,ins,talent,Bs1,Bs2):
    if isNotFull(band):
        if checkGender(band_gender,gender):
            if checkInstrument(band,ins):
                if checkTalent(band,talent):
                    return True
                    '''if checkBandStatus(Bs1,Bs2,band,ins,talent):
                        return True
                    else:
                        return False'''
                else:
                    return False
            else: return False
        else:
            return False
    else:
        return False

def isNotFull(band):
    if countIns(band) < 6:
        return True
    else:
        return False
        
def countIns(band):
    count = 0
    for i in band.keys():
        if  i in ['S','s','D','d','K','k','I','i','G','g','B','b']:
            count += 1
    return count

def checkInstrument(band,ins):
    if ins not in band.keys():
        return True
    else:
        return False

def checkTalent(band,talent):
    talentList = [countTalent(band,1),countTalent(band,2),countTalent(band,3),countTalent(band,4)]
    if talent == 1:
        if talentList[0] == 0:
            return True
        elif talentList[0] == 1:
            if band['type'] == '1':
                return True
            else: 
                return False

        elif talentList[0] == 2:
            return False
    if talent == 2:
        if talentList[1] == 0:
            return True
        elif talentList[1] == 1:
            if band['type'] == '2':
                return True
            else: 
                return False

        elif talentList[1] == 2:
            return False
    if talent == 3:
        if talentList[2] == 0:
            return True
        elif talentList[2] == 1:
            if band['type'] == '2':
                return True
            else: 
                return False

        elif talentList[2] == 2:
            return False
    if talent == 4:
        if talentList[3] == 0:
            return True
        elif talentList[3] == 1:
            if band['type'] == '1':
                return True
            else: 
                return False
        elif talentList[3] == 2:
            return False


def checkGender(band_gender,gender):
    if countGender(band_gender,gender) < 3:
        return True
    else:
        return False

def countTalent(band,talent):
    result = countIf(band.values(),talent)
    return result

def countGender(band_gender,gender):
    count = 0
    for g in band_gender:
        if g == gender:
            count += 1
    return count

def sumOfTalent(band):
    result = 0
    for i in band.keys():
        if i in ['S','s','D','d','K','k','I','i','G','g','B','b']:
            result = result + int(band[i])
    return result

def countIf(list,x):
    count = 0
    for i in list:
        if x == i:
            count += 1
    return count



def talentStatus(b1,b2,b3,b4,ins,talent):
    b_total = []
    if checkInstrument(b1,ins):
        if checkTalent(b1,talent):
            b_total.append(b1)
    if checkInstrument(b2,ins):
        if checkTalent(b2,talent):
            b_total.append(b2)
    if checkInstrument(b3,ins):
        if checkTalent(b3,talent):
            b_total.append(b3)
    if checkInstrument(b4,ins):
        if checkTalent(b4,talent):
            b_total.append(b4)
    return b_total

def ageChoice(b_total,age):
    def_age_list = buildAgeList(b_total,age)
    for i in range(len(def_age_list)):
        if def_age_list[i] == max(def_age_list):
            return b_total[i]['ID']

def buildAgeList(b_total,age):
    age_list = []
    for b in b_total:
        age_list.append( abs(avg(b['age'])- age ) )
    return age_list

def avg(band_age):
    if len(band_age) == 0:
        return 0
    else:
        return sum(band_age)/len(band_age)
