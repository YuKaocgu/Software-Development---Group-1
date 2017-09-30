import random
import sys
import os
import numpy as np
import openpyxl as px
import pandas as pd
from pandasql import sqldf


 # Fill in Data .
wb=px.load_workbook('C:\DD.xlsx')
ws=wb.get_sheet_by_name("Sheet1")

#xl = pd.read_excel(r'C:\DD.xls',"Sheet1",0)

    #read_excel(r'C:\DD.xlsx',"Sheet1",0)
pysqldf = lambda q: sqldf(q, globals())
nRow = 2

nlRow = ws.max_row+1
nlColumn = ws.max_column
# Capturing each row of data from columns.
for row in range(nRow, nlRow):
   # Only targeting Accepted candidates
   status = ws['J' + str(row)].value
   if status=="Accepted":
    camp = ws['C' + str(row)].value


    for band in ["Group1","Group2","Group3","Group4","Group5","Group6","Group7","Group8"]:
       continue

    #June Selection
    if camp == "June":
        gender = list(ws['E' + str(row)].value)
        name = ws['D' + str(row)].value
        band = ws['L' + str(row)].value
        talent = ws['M' + str(row)].value
        instrument = ws['N' + str(row)].value
        git = ws['P' + str(row)].value
        '''
        Need help in writing the data to excel, then counting and summing the values for further analysis. 
        Count, Append, Sum 
        '''
        band.append("Group1")

        a = band.count()
        band.sum()
        print(name, band, a)


'''

    # July Selection
    elif camp == "July":
        gender = ws['E' + str(row)].value
        name = ws['D' + str(row)].value
        band = ws['L' + str(row)].value
        talent = ws['M' + str(row)].value
        instrument = ws['N' + str(row)].value
        git = ws['P' + str(row)].value
        print("They start in July")



    # August Selection
    elif camp == "August":
        gender = ws['E' + str(row)].value
        name = ws['D' + str(row)].value
        band = ws['L' + str(row)].value
        talent = ws['M' + str(row)].value
        instrument = ws['N' + str(row)].value
        git = ws['P' + str(row)].value
        print("They start in August")


    '''
#if count1 < 7:
 #           band= "Group1"
  #      else:
   #         band = "Group2"
    #    print(name,band)
   # '''

'''
#pysqldf = lambda q: sqldf(q, globals())
#print (pysqldf("Select [Name],[Gender],[Talent],[Instrument] from xl where [Decision]='Accepted'"))

'''