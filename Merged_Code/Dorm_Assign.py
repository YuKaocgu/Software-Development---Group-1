import numpy
import openpyxl as px
import pandas as pd
import warnings

def test():
    age = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24]
    Dorm1 = []
    Dorm2 = []
    Dorm3 = []
    for i in age:
        result = dorm_assign(Dorm1,Dorm2,Dorm3,i)
        print(result)
def isFull(dorm):
    if (len(dorm) == 8):
        return True
    else:
        return False
def avg(data_List):
    if len(data_List) == 0:
        return 0
    else:
        return sum(data_List)/len(data_List)
def dorm_assign(Dorm1,Dorm2,Dorm3,age):
    Dorm_Status = [isFull(Dorm1),isFull(Dorm2),isFull(Dorm3)]
    if Dorm_Status == [False,False,False]:
        check_list = [ abs(avg(Dorm1)-age),abs(avg(Dorm2)-age),abs(avg(Dorm3)-age)]
        if max(check_list) == abs(avg(Dorm1)-age):
            Dorm1.append(age)
            return 1
        elif max(check_list) == abs(avg(Dorm2)-age):
            Dorm2.append(age)
            return 2
        else:
            Dorm3.append(age)
            return 3
    elif Dorm_Status == [True,False,False]:
        if abs(avg(Dorm2)-age) > abs(avg(Dorm3)-age):
            Dorm2.append(age)
            return 2
        else:
            Dorm3.append(age) 
            return 3
    elif Dorm_Status == [False,True,False]:
        if abs(avg(Dorm1)-age) > abs(avg(Dorm3)-age):
            Dorm1.append(age)
            return 1
        else:
            Dorm3.append(age) 
            return 3
    elif Dorm_Status == [False,False,True]:
        if abs(avg(Dorm1)-age) > abs(avg(Dorm2)-age):
            Dorm1.append(age)
            return 1
        else:
            Dorm2.append(age) 
            return 2
    elif Dorm_Status == [False,True,True]:
        Dorm1.append(age)
        return 1
    elif Dorm_Status == [True,False,True]:
        Dorm2.append(age)
        return 2
    elif Dorm_Status == [True,True,False]:
        Dorm3.append(age)
        return 3

def start_assign():
    url = "DD.xlsx"
     # Fill in Data .
    warnings.simplefilter("ignore") 
    wb=px.load_workbook(url)
    ws=wb.get_sheet_by_name("Sheet1")
    sheet = wb.active
    DormM1 = []
    DormM2 = []
    DormM3 = []
    DormF1 = []
    DormF2 = []
    DormF3 = []
    
    nRow = 2
    nlRow = ws.max_row+1
    nlColumn = ws.max_column
    # Capturing each row of data from columns.
    for row in range(nRow, nlRow):
        # Only targeting Accepted candidates
        status = ws['L' + str(row)].value
        if status=="A":
            camp = ws['H' + str(row)].value
            Fname = ws['B' + str(row)].value
            Lname = ws['C' + str(row)].value
            gender = (ws['E' + str(row)].value)
            age = int((ws['F' + str(row)].value))
            dorm = str(ws['O' + str(row)].value)
            #Dorm assign:
            if gender == 'M':
                dorm_result = Dorm_Assign.dorm_assign(DormM1,DormM2,DormM3,age)
                if dorm_result == 1 :
                    sheet['O' + str(row)] = "M1"
                elif dorm_result == 2:
                    sheet['O' + str(row)] = "M2"
                else:
                    sheet['O' + str(row)] = "M3"
            if gender == 'F':
                dorm_result = Dorm_Assign.dorm_assign(DormF1,DormF2,DormF3,age)
                if dorm_result == 1 :
                    sheet['O' + str(row)] = "F1"
                elif dorm_result == 2:
                    sheet['O' + str(row)] = "F2"
                else:
                    sheet['O' + str(row)] = "F3"
    wb.save(url)
