#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Oct  8 17:24:59 2017

@author: ChristofferGamborg
"""
import openpyxl as px
import pandas as pd
def CheckIn():
    url = "DD.xlsx"
    wb=px.load_workbook(url)
    ws=wb.get_sheet_by_name("Sheet1")
    sheet = wb.active
    nRow = 2
    nlRow = ws.max_row+1
    while True:
        Choose = input ('\n'.join(['1: Check by name', '2: Check by IDNumber', '3: Return to Main Menu' , '']))
        while Choose not in ['1','2', '3']:
            print('Invalid input')
            Choose = input ('\n'.join(['1: Check by name', '2: Check by IDNumber']))
        if Choose == '1':    
            StudentFname = input ('Write the student first name: \n')
            StudentLname = input ('Write the student last name: \n')
            Checking_Status = False
            for row in range(nRow, nlRow):
                # Only targeting Accepted candidates
                Fname = ws['B' + str(row)].value
                Lname = ws['C' + str(row)].value
                if Fname == StudentFname and Lname == StudentLname:
                        print("Found \n")
                        CheckIn = input ('Did ' + Fname + " " + Lname + 'meet all of the requirements to Check In? (Y or N) \n' )
                        while CheckIn not in ['Y','y','n','N']:
                            print ('Invalid Input')
                            CheckIn = input ('Did ' + Fname + " " + Lname + ' meet all of the requirements to Check In? (Yes or No) \n' )
                        if CheckIn in ['Y', 'y']:
                            print(Fname + ' ' + Lname + ' is Checked In')
                            sheet['Q' + str(row)] = 'Y'
                            wb.save(url)
                            Checking_Status = True
                        elif CheckIn in ['N', 'n']:
                            print ('Sent him home, his an idiot')
                            sheet['Q' + str(row)] = 'N'
                            wb.save(url)
                else:
                    continue
            
        if Choose == '2':
            StudentID = input("Please insert ID number: \n")
            for row in range(nRow, nlRow):
                ID = ws['A' + str(row)].value
                if ID == StudentID:
                    Fname = ws['B' + str(row)].value
                    Lname = ws['C' + str(row)].value
                    print(Fname + " " + Lname)
                    respond = input("Is this you? (Y or N")
                    if respond in ['Y', 'y']:
                        print("Found")
                        CheckIn = input ('Did ' + Fname + ' ' + Lname + 'meet all of the requirements to Check In? (Yes or No) \n' )
                        while CheckIn not in ['Y','y','n','N']:
                            print ('Invalid Input')
                            CheckIn = input ('Did ' + Fname + ' ' + Lname + 'meet all of the requirements to Check In? (Yes or No) \n' )
                            if CheckIn in ['Y', 'y']:
                                sheet['Q' + str(row)] = 'Y'
                                wb.save(url)
                                Checking_Status = True
                            elif CheckIn in ['N', 'n']:
                                sheet['Q' + str(row)] = 'N'
                                wb.save(url)
        if Choose == '3':
            break

        if Checking_Status:
            print ("XXX checked")
            continue
        else:
            print("Not in Data")
            continue
            
                        
       
       