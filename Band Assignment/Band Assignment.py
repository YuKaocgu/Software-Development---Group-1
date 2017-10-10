import openpyxl as px
import pandas as pd

url = "G:\Master - FE & IST\Information Systems And Technology\Software Development And Programming\Project\Algorithim\DD.xlsx"
 # Fill in Data .
wb=px.load_workbook(url)
ws=wb.get_sheet_by_name("Sheet1")
sheet = wb.active



group1_count = group2_count = group3_count = group4_count = group5_count = group6_count = group7_count = group8_count = 0
gm1_count = gf1_count = gm2_count = gf2_count = gm3_count = gf3_count =gm4_count = gf4_count = gm5_count = gf5_count = gm6_count = gf6_count = gm7_count = gf7_count = gm8_count = gf8_count = 0
s1 = g1 = k1 = b1 = d1 = i1 = s2 = g2 = k2 = b2 = d2 = i2 = s3 = g3 = k3 = b3 = d3 = i3 = s4 = g4 = k4 = b4 = d4 = i4 = s5 = g5 = k5 = b5 = d5 = i5 = s6 = g6 = k6 = b6 = d6 = i6 = s7 = g7 = k7 = b7 = d7 = i7 = s8 = g8 = k8 = b8 = d8 = i8 = 0
t1 = t2 = t3 =t4 =t5=t6=t21=t22=t23=t24=t25=t26 = t31=t32=t33=t34=t35=t36= t51=t52=t53=t54=t55=t56=t61=t62=t63=t64=t65=t66=t71=t72=t73=t74=t75=t76=t81=t82=t83=t84=t85=t86=t41=t42=t43=t44=t45=t46=0


nRow = 1
nlRow = ws.max_row+1
nlColumn = ws.max_column
# Capturing each row of data from columns.
for row in range(nRow, nlRow):
   # Only targeting Accepted candidates
   status = ws['J' + str(row)].value
   if status=="Accepted":
    camp = ws['C' + str(row)].value
    name = ws['D' + str(row)].value
    gender = (ws['E' + str(row)].value)
    talent = ws['M' + str(row)].value
    instrument = ws['N' + str(row)].value
    band = str(ws['L' + str(row)].value)

    # Group 1 assignment
    if band == "None":
        if talent == 1 and t1 < 1 :
            if group1_count < 6:
                if s1<1 and gender in ('M','F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M','F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t1 +=1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t1 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
        elif talent == 2 and t2<1:
            if group1_count < 6:
                if s1<1 and gender in ('M','F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M','F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t2 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
        elif talent == 3 and t3 < 1:
            if group1_count < 6:
                if s1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t3 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
        elif talent == 4 and t4 < 1:
            if group1_count < 6:
                if s1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t4 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
        elif talent == 1 and t5 < 1:
            if group1_count < 6:
                if s1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t5 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                           
        elif talent == 4 and t6 < 1:
            if group1_count < 6:
                if s1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            s1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            s1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if k1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            k1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group1"
                            gf1_count += 1
                            group1_count += 1
                            k1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if i1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            i1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if d1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            d1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if g1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            g1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
            if group1_count < 6:
                if b1 < 1 and gender in ('M', 'F'):
                    if gm1_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
                    if gf1_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group1"
                            gm1_count += 1
                            group1_count += 1
                            b1 += 1
                            t6 += 1
                            sheet['L' + str(row)] = 1
                            wb.save(url)
    
        # Group 2 assignment
    if band == "None":
        if talent == 2 and t21 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t21 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
        elif talent == 3 and t22 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t22 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
        elif talent == 4 and t23 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t23 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
        elif talent == 1 and t24 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t24 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
        elif talent == 2 and t25 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t25 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)

        elif talent == 3 and t26 < 1:
            if group2_count < 6:
                if s2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            s2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            s2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if k2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            k2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group2"
                            gf2_count += 1
                            group2_count += 1
                            k2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if i2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            i2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if d2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            d2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if g2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            g2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
            if group2_count < 6:
                if b2 < 1 and gender in ('M', 'F'):
                    if gm2_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
                    if gf2_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group2"
                            gm2_count += 1
                            group2_count += 1
                            b2 += 1
                            t26 += 1
                            sheet['L' + str(row)] = 2
                            wb.save(url)
    # Group 3 assignment
    if band == "None":
        if talent == 3 and t31 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t31 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
        elif talent == 4 and t32 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t32 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
        elif talent == 1 and t33 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t33 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
        elif talent == 2 and t34 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t34 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
        elif talent == 4 and t35 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t35 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)

        elif talent == 1 and t36 < 1:
            if group3_count < 6:
                if s3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            s3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            s3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if k3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            k3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group3"
                            gf3_count += 1
                            group3_count += 1
                            k3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if i3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            i3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if d3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            d3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if g3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            g3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
            if group3_count < 6:
                if b3 < 1 and gender in ('M', 'F'):
                    if gm3_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)
                    if gf3_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group3"
                            gm3_count += 1
                            group3_count += 1
                            b3 += 1
                            t36 += 1
                            sheet['L' + str(row)] = 3
                            wb.save(url)

    # Group 4 assignment
    if band == "None":
        if talent == 4 and t41 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t41 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
        elif talent == 1 and t42 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t42 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
        elif talent == 2 and t43 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t43 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
        elif talent == 3 and t44 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t44 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
        elif talent == 3 and t45 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t45 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)

        elif talent == 2 and t46 < 1:
            if group4_count < 6:
                if s4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            s4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            s4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if k4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            k4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group4"
                            gf4_count += 1
                            group4_count += 1
                            k4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if i4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            i4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if d4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            d4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if g4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            g4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
            if group4_count < 6:
                if b4 < 1 and gender in ('M', 'F'):
                    if gm4_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)
                    if gf4_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group4"
                            gm4_count += 1
                            group4_count += 1
                            b4 += 1
                            t46 += 1
                            sheet['L' + str(row)] = 4
                            wb.save(url)

        # Group 5 assignment
    if band == "None":
        if talent == 1 and t51 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t51 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
        elif talent == 2 and t52 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t52 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
        elif talent == 3 and t53 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t53 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
        elif talent == 4 and t54 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t54 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
        elif talent == 1 and t55 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t55 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)

        elif talent == 4 and t56 < 1:
            if group5_count < 6:
                if s5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            s5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            s5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if k5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            k5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group5"
                            gf5_count += 1
                            group5_count += 1
                            k5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if i5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            i5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if d5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            d5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if g5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            g5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
            if group5_count < 6:
                if b5 < 1 and gender in ('M', 'F'):
                    if gm5_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)
                    if gf5_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group5"
                            gm5_count += 1
                            group5_count += 1
                            b5 += 1
                            t56 += 1
                            sheet['L' + str(row)] = 5
                            wb.save(url)

        # Group 6 assignment
    if band == "None":
        if talent == 2 and t61 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t61 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
        elif talent == 3 and t62 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t62 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
        elif talent == 4 and t63 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t63 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
        elif talent == 1 and t64 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t64 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
        elif talent == 2 and t65 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t65 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)

        elif talent == 3 and t66 < 1:
            if group6_count < 6:
                if s6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            s6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            s6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if k6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            k6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group6"
                            gf6_count += 1
                            group6_count += 1
                            k6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if i6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            i6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if d6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            d6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if g6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            g6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
            if group6_count < 6:
                if b6 < 1 and gender in ('M', 'F'):
                    if gm6_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
                    if gf6_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group6"
                            gm6_count += 1
                            group6_count += 1
                            b6 += 1
                            t66 += 1
                            sheet['L' + str(row)] = 6
                            wb.save(url)
        # Group 7 assignment
    if band == "None":
        if talent == 3 and t71 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t71 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
        elif talent == 4 and t72 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t72 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
        elif talent == 1 and t73 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t73 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
        elif talent == 2 and t74 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t74 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
        elif talent == 4 and t75 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t75 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
    
        elif talent == 1 and t76 < 1:
            if group7_count < 6:
                if s7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            s7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            s7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if k7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            k7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group7"
                            gf7_count += 1
                            group7_count += 1
                            k7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if i7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            i7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if d7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            d7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if g7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            g7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
            if group7_count < 6:
                if b7 < 1 and gender in ('M', 'F'):
                    if gm7_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)
                    if gf7_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group7"
                            gm7_count += 1
                            group7_count += 1
                            b7 += 1
                            t76 += 1
                            sheet['L' + str(row)] = 7
                            wb.save(url)

    # Group 8 assignment
    if band == "None":
        if talent == 4 and t81 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t81 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
        elif talent == 1 and t82 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t82 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
        elif talent == 2 and t83 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t83 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
        elif talent == 3 and t84 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t84 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
        elif talent == 3 and t85 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t85 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)

        elif talent == 2 and t86 < 1:
            if group8_count < 6:
                if s8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Singer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            s8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Singer':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            s8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if k8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Keyboardist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            k8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Keyboardist':
                            band = "Group8"
                            gf8_count += 1
                            group8_count += 1
                            k8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if i8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Instrumentalist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            i8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if d8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Drummer':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            d8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if g8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Guitarist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            g8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
            if group8_count < 6:
                if b8 < 1 and gender in ('M', 'F'):
                    if gm8_count < 3:
                        if gender == 'M' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
                    if gf8_count < 3:
                        if gender == 'F' and instrument == 'Bassist':
                            band = "Group8"
                            gm8_count += 1
                            group8_count += 1
                            b8 += 1
                            t86 += 1
                            sheet['L' + str(row)] = 8
                            wb.save(url)
    print (name, talent, instrument, gender, band)
