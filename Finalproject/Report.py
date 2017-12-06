
import pandas as pd
from pandas import ExcelWriter
from pandasql import sqldf
import openpyxl as px
import xlsxwriter

url = "DD.xlsx"
xl = pd.read_excel(url)
# Fill in Data .
wb = px.load_workbook(url)
workbook = xlsxwriter.Workbook('DD.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
wm1
sheet = wb.active
M1 = workbook.add_worksheet()

NewList = {}
M1 = M2 = M3 = F1 = F2 = F3 = {}


def Dorm(row):
    status = ws['L' + str(row)].value
    if status in ["A","a"]:
        realDorm = (ws['O' + str(row)].value)
        if realDorm in ["M1","m1"]:
            Id = ws['A' + str(row)].value
            FName=ws['B' + str(row)].value
            LName=ws['C' + str(row)].value
            Camp = ws['H' + str(row)].value
            ListOfInformation = [Id,FName,LName,Camp,realDorm]
            M1[(Id)] = ListOfInformation
            
            
            
        

def report_dorm_assign():
    url = "DD.xlsx"
     # Fill in Data .
    wb=px.load_workbook(url)
    ws=wb.get_sheet_by_name("Sheet1")
    sheet = wb.active
    
    nRow = 1
    nlRow = ws.max_row+1
    # Capturing each row of data from columns.
    for row in range(nRow, nlRow):
        Dorm(row)
                
    


    
    """url = "DD.xlsx"
    sav = ExcelWriter("Band Assignment.xls")

    xl = pd.read_excel(url,"Sheet1",0)
    pysqldf = lambda q: sqldf(q, globals())

    M1 = pysqldf("SELECT [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='M1' order by [First Name]")
    M1.to_excel(sav, sheet_name='Male Dorm1', index=False)

    M2 = pysqldf("Select [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='M2' order by [First Name]")
    M2.to_excel(sav, sheet_name='Male Dorm2', index=False)

    M3 = pysqldf("Select [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='M3' order by [First Name]")
    M3.to_excel(sav, sheet_name='Male Dorm3', index=False)

    F1 = pysqldf("Select [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='F1' order by [First Name]")
    F1.to_excel(sav, sheet_name='Female Dorm1', index=False)

    F2 = pysqldf("Select [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='F2' order by [First Name]")
    F2.to_excel(sav, sheet_name='Female Dorm2', index=False)

    F3 = pysqldf("Select [ID Number],[First Name], [Last Name], [Dorm] from xl where [Status]='A' and [Dorm]='F3' order by [First Name]")
    F3.to_excel(sav, sheet_name='Female Dorm3', index=False)

    DE = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A' and [Dorm]='' order by [First Name]")
    DE.to_excel(sav, sheet_name='Unassigned', index=False)

    sav.save()


def report_band_assign():

    url = "DD.xlsx"
    sav = ExcelWriter("Band Assignment.xls")

    xl = pd.read_excel(url,"Sheet1",0)
    pysqldf = lambda q: sqldf(q, globals())


    B1 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='1' order by [First Name]")
    B1.to_excel(sav, sheet_name='Band1', index=False)


    B2 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='2' order by [First Name]")
    B2.to_excel(sav, sheet_name='Band2', index=False)

    B3 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='3' order by [First Name]")
    B3.to_excel(sav, sheet_name='Band3', index=False)

    B4 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='4' order by [First Name]")
    B4.to_excel(sav, sheet_name='Band4', index=False)

    B5 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='5' order by [First Name]")
    B5.to_excel(sav, sheet_name='Band5', index=False)

    B6 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='6' order by [First Name]")
    B6.to_excel(sav, sheet_name='Band6', index=False)

    B7 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='7' order by [First Name]")
    B7.to_excel(sav, sheet_name='Band7', index=False)

    B8 = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='8' order by [First Name]")
    B8.to_excel(sav, sheet_name='Band8', index=False)

    BE = pysqldf("Select [ID Number],[First Name], [Last Name], [Band] from xl where [Status]='A'  and [Band]='' order by [First Name]")
    BE.to_excel(sav, sheet_name='Unassigned', index=False)

    sav.save()
"""